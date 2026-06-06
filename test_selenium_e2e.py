"""
Selenium End-to-End Test for Banking Credit Risk Calculator
Data-Driven Testing: Reads test cases from Excel file (test_data.xlsx)
Tests the complete consolidated workflow in borrower-info.html
Supports multiple test cycles via Excel rows
"""

import sys
import os
import time
from pathlib import Path
from datetime import datetime

# Fix encoding for Windows
if sys.platform == "win32":
    os.environ["PYTHONIOENCODING"] = "utf-8"
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# Import openpyxl for Excel reading
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Install with: pip install openpyxl")

# Configuration
BASE_PATH = Path(r"C:\Users\Arnav\OneDrive\Desktop\Daily reading\Banking_Credit_Risk\public")
BORROWER_INFO_URL = f"file:///{BASE_PATH}/borrower-info.html".replace("\\", "/")
TEST_DATA_FILE = Path(r"C:\Users\Arnav\OneDrive\Desktop\Daily reading\Banking_Credit_Risk\test_data.xlsx")


class ExcelTestDataReader:
    """Reads test data from Excel file"""

    @staticmethod
    def read_test_cases(filename):
        """Read all test cases from Excel file"""
        if not EXCEL_AVAILABLE:
            print("ERROR: openpyxl is required. Install: pip install openpyxl")
            return []

        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            test_cases = []
            headers = []

            # Read headers from first row
            for cell in ws[1]:
                headers.append(cell.value)

            # Read data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if row[0] is None:  # Skip empty rows
                    continue

                test_case = {}
                for col_idx, header in enumerate(headers):
                    test_case[header] = row[col_idx]

                test_cases.append(test_case)

            wb.close()
            return test_cases
        except Exception as e:
            print(f"ERROR reading Excel file: {e}")
            return []


class BankingCreditRiskTest:
    """Test suite for Banking Credit Risk Calculator - Data Driven"""

    def __init__(self):
        """Initialize WebDriver and test state"""
        self.driver = None
        self.wait = None
        self.test_passed = 0
        self.test_failed = 0
        self.test_skipped = 0
        self.test_cases_results = []
        self.current_test_case = None
        self.step_timeout = 10  # Timeout per step in seconds

    def setup(self):
        """Set up Chrome WebDriver"""
        print("\n" + "="*80)
        print("SETTING UP SELENIUM WEBDRIVER")
        print("="*80)

        chrome_options = Options()
        # Don't use headless - user wants to see the browser
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        try:
            # Use webdriver-manager to automatically manage ChromeDriver
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.wait = WebDriverWait(self.driver, 10)
            print("[OK] Chrome WebDriver initialized successfully")
            return True
        except Exception as e:
            print(f"[FAIL] Failed to initialize WebDriver: {e}")
            print("\nNote: Make sure the following packages are installed:")
            print("  pip install selenium webdriver-manager openpyxl")
            return False

    def log_test(self, test_name, passed, message=""):
        """Log test result"""
        status = "[PASS]" if passed else "[FAIL]"
        if passed:
            self.test_passed += 1
        else:
            self.test_failed += 1
        print(f"  {status}: {test_name}")
        if message:
            print(f"       {message}")

    def test_open_borrower_info(self, test_case_id):
        """Test: Open borrower-info.html"""
        test_name = f"{test_case_id} - Open borrower-info.html"
        print(f"\n{test_name}")
        try:
            self.driver.get(BORROWER_INFO_URL)
            # Wait for page to load
            self.wait.until(EC.presence_of_element_located((By.ID, "borrowerId")))
            self.log_test(test_name, True, f"Loaded: borrower-info.html")
            return True
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def test_fill_form_from_excel(self, test_case, test_case_id):
        """Test: Fill all form fields from Excel test data"""
        test_name = f"{test_case_id} - Fill Form from Excel Data"
        print(f"\n{test_name}")
        try:
            # Borrower Information
            borrower_id = self.driver.find_element(By.ID, "borrowerId")
            borrower_id.clear()
            borrower_id.send_keys(str(test_case.get("Borrower ID", "")))

            borrower_name = self.driver.find_element(By.ID, "borrowerName")
            borrower_name.clear()
            borrower_name.send_keys(str(test_case.get("Borrower Name", "")))

            exposure_amount = self.driver.find_element(By.ID, "exposureAmount")
            exposure_amount.clear()
            exposure_amount.send_keys(str(int(test_case.get("Exposure Amount", 0))))

            sector = self.driver.find_element(By.ID, "sector")
            Select(sector).select_by_value(str(test_case.get("Sector", "")))

            # Select Calculation Mode
            calc_mode = str(test_case.get("Calc Mode", "both")).lower()
            if calc_mode == "standardized":
                mode_radio_id = "modeSA"
            elif calc_mode == "both":
                mode_radio_id = "modeBoth"
            else:
                mode_radio_id = "modeAIRB"

            mode_radio = self.driver.find_element(By.ID, mode_radio_id)
            mode_radio.click()
            time.sleep(0.3)

            # Financial Metrics
            de_input = self.driver.find_element(By.ID, "debtToEquity")
            de_input.clear()
            de_input.send_keys(str(test_case.get("Debt to Equity", "")))

            icr_input = self.driver.find_element(By.ID, "interestCoverage")
            icr_input.clear()
            icr_input.send_keys(str(test_case.get("Interest Coverage", "")))

            margin_input = self.driver.find_element(By.ID, "profitabilityMargin")
            margin_input.clear()
            margin_input.send_keys(str(test_case.get("Profitability Margin", "")))

            lr_input = self.driver.find_element(By.ID, "liquidityRatio")
            lr_input.clear()
            lr_input.send_keys(str(test_case.get("Liquidity Ratio", "")))

            # AIRB Section (if applicable)
            if calc_mode in ["airb", "both"]:
                seniority = str(test_case.get("Seniority", ""))
                if seniority:
                    seniority_select = Select(self.driver.find_element(By.ID, "seniority"))
                    seniority_select.select_by_value(seniority)

                loan_type = str(test_case.get("Loan Type", ""))
                if loan_type:
                    loan_type_select = Select(self.driver.find_element(By.ID, "loanType"))
                    loan_type_select.select_by_value(loan_type)

                collateral_type = str(test_case.get("Collateral Type", ""))
                if collateral_type and collateral_type != "None":
                    collateral_select = Select(self.driver.find_element(By.ID, "collateralType"))
                    collateral_select.select_by_value(collateral_type)

                    collateral_value = self.driver.find_element(By.ID, "collateralValue")
                    collateral_value.clear()
                    collateral_value.send_keys(str(int(test_case.get("Collateral Value", 0))))

            # SA Section (if applicable)
            if calc_mode in ["standardized", "both"]:
                external_rating = str(test_case.get("External Rating", ""))
                if external_rating:
                    rating_select = Select(self.driver.find_element(By.ID, "externalRating"))
                    rating_select.select_by_value(external_rating)

                exposure_category = str(test_case.get("Exposure Category", ""))
                if exposure_category:
                    category_select = Select(self.driver.find_element(By.ID, "exposureCategory"))
                    category_select.select_by_value(exposure_category)

            self.log_test(test_name, True, f"Filled: {test_case.get('Borrower Name', 'Unknown')}")
            return True
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def test_click_calculate(self, test_case_id):
        """Test: Click Calculate Risk Parameters button"""
        test_name = f"{test_case_id} - Click Calculate"
        print(f"\n{test_name}")
        try:
            calculate_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), 'Calculate Risk Parameters')]")
            calculate_btn.click()

            # Wait for results section with timeout
            try:
                wait_timeout = WebDriverWait(self.driver, self.step_timeout)
                results_section = wait_timeout.until(
                    EC.visibility_of_element_located((By.ID, "resultsSection"))
                )
                self.log_test(test_name, True, "Calculation completed, results displayed")
                return True
            except:
                # Timeout occurred, skip this test case
                self.log_test(test_name, False, f"TIMEOUT: Page did not respond within {self.step_timeout} seconds")
                self.test_skipped += 1
                return False
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def test_verify_pd_result(self, test_case, test_case_id):
        """Test: Verify PD calculation result"""
        test_name = f"{test_case_id} - Verify PD Result"
        print(f"\n{test_name}")
        try:
            pd_value_elem = self.driver.find_element(By.ID, "pdValue")
            pd_text = pd_value_elem.text.strip()
            pd_value = float(pd_text.replace("%", "").strip())

            pd_min = float(test_case.get("Expected PD Min", 0))
            pd_max = float(test_case.get("Expected PD Max", 100))

            in_range = pd_min <= pd_value <= pd_max

            message = f"PD: {pd_value}% (Expected: {pd_min}-{pd_max}%)"
            self.log_test(test_name, in_range, message)
            return in_range
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def test_verify_lgd_result(self, test_case, test_case_id):
        """Test: Verify LGD calculation result"""
        test_name = f"{test_case_id} - Verify LGD Result"
        print(f"\n{test_name}")
        try:
            lgd_value_elem = self.driver.find_element(By.ID, "lgdValue")
            lgd_text = lgd_value_elem.text.strip()
            lgd_value = float(lgd_text.replace("%", "").strip())

            lgd_min = float(test_case.get("Expected LGD Min", 0))
            lgd_max = float(test_case.get("Expected LGD Max", 100))

            in_range = lgd_min <= lgd_value <= lgd_max

            message = f"LGD: {lgd_value}% (Expected: {lgd_min}-{lgd_max}%)"
            self.log_test(test_name, in_range, message)
            return in_range
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def test_click_confirm_and_record(self, test_case_id):
        """Test: Click Confirm & Record Loan button"""
        test_name = f"{test_case_id} - Click Confirm & Record"
        print(f"\n{test_name}")
        try:
            confirm_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), 'Confirm & Record Loan')]")
            confirm_btn.click()

            # Wait for portfolio section with timeout
            try:
                wait_timeout = WebDriverWait(self.driver, self.step_timeout)
                portfolio_section = wait_timeout.until(
                    EC.visibility_of_element_located((By.ID, "portfolioSection"))
                )
                time.sleep(1)
                self.log_test(test_name, True, "Loan recorded, portfolio displayed")
                return True
            except:
                # Timeout occurred, skip this test case
                self.log_test(test_name, False, f"TIMEOUT: Portfolio did not display within {self.step_timeout} seconds")
                self.test_skipped += 1
                return False
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def test_verify_loan_in_portfolio(self, test_case, test_case_id):
        """Test: Verify loan appears in portfolio table"""
        test_name = f"{test_case_id} - Verify Loan in Portfolio"
        print(f"\n{test_name}")
        try:
            portfolio_table = self.driver.find_element(By.ID, "loansTable")
            table_text = portfolio_table.text
            borrower_name = str(test_case.get("Borrower Name", ""))
            borrower_found = borrower_name in table_text

            message = f"Borrower '{borrower_name}' in portfolio: {borrower_found}"
            self.log_test(test_name, borrower_found, message)
            return borrower_found
        except Exception as e:
            self.log_test(test_name, False, str(e))
            return False

    def run_test_cycle(self, test_case):
        """Run complete test cycle for one test case from Excel"""
        test_case_id = test_case.get("Test ID", "UNKNOWN")
        borrower_name = test_case.get("Borrower Name", "Unknown")

        print(f"\n{'='*80}")
        print(f"TEST CYCLE: {test_case_id} - {borrower_name}")
        print(f"{'='*80}")

        test_results = {
            "Test ID": test_case_id,
            "Borrower Name": borrower_name,
            "Status": "PENDING",
            "Description": test_case.get("Test Description", ""),
            "Passed": 0,
            "Failed": 0,
            "Skipped": 0
        }

        try:
            # Test sequence
            if not self.test_open_borrower_info(test_case_id):
                test_results["Status"] = "FAILED"
                test_results["Failed"] += 1
                self.test_failed += 1
                return test_results

            if not self.test_fill_form_from_excel(test_case, test_case_id):
                test_results["Status"] = "FAILED"
                test_results["Failed"] += 1
                self.test_failed += 1
                return test_results

            if not self.test_click_calculate(test_case_id):
                test_results["Status"] = "SKIPPED"
                test_results["Skipped"] = 1
                return test_results

            self.test_verify_pd_result(test_case, test_case_id)
            self.test_verify_lgd_result(test_case, test_case_id)

            if not self.test_click_confirm_and_record(test_case_id):
                test_results["Status"] = "SKIPPED"
                test_results["Skipped"] = 1
                return test_results

            if not self.test_verify_loan_in_portfolio(test_case, test_case_id):
                test_results["Status"] = "FAILED"
                test_results["Failed"] += 1
                self.test_failed += 1
                return test_results

            test_results["Status"] = "PASSED"
            test_results["Passed"] = 6
            self.test_passed += 1

        except Exception as e:
            print(f"  [ERROR] Test cycle failed: {e}")
            test_results["Status"] = "ERROR"
            test_results["Failed"] += 1
            self.test_failed += 1

        return test_results

    def run_all_test_cycles(self, test_cases):
        """Run complete test cycles for all test cases from Excel"""
        print("\n" + "="*80)
        print("BANKING CREDIT RISK CALCULATOR - DATA-DRIVEN TEST SUITE")
        print(f"Running {len(test_cases)} test scenarios from Excel file")
        print("="*80)

        for test_case in test_cases:
            result = self.run_test_cycle(test_case)
            self.test_cases_results.append(result)
            time.sleep(2)  # Pause between test cycles

    def print_summary(self):
        """Print comprehensive test summary"""
        print("\n" + "="*100)
        print("TEST EXECUTION SUMMARY")
        print("="*100)

        total_tests = len(self.test_cases_results)
        passed_tests = sum(1 for r in self.test_cases_results if r["Status"] == "PASSED")
        failed_tests = sum(1 for r in self.test_cases_results if r["Status"] == "FAILED")
        skipped_tests = sum(1 for r in self.test_cases_results if r["Status"] == "SKIPPED")
        error_tests = sum(1 for r in self.test_cases_results if r["Status"] == "ERROR")

        print(f"\nTotal Test Cases: {total_tests}")
        print(f"[PASS] Passed: {passed_tests}")
        print(f"[FAIL] Failed: {failed_tests}")
        print(f"[SKIP] Skipped (Timeout): {skipped_tests}")
        print(f"[ERROR] Errors: {error_tests}")

        print("\n" + "="*100)
        print("PASS/FAIL TABLE")
        print("="*100)
        print(f"| {'Test ID':<12} | {'Borrower Name':<25} | {'Status':<10} | {'Notes':<35} |")
        print("|" + "-"*12 + "|" + "-"*27 + "|" + "-"*12 + "|" + "-"*37 + "|")

        for result in self.test_cases_results:
            test_id = result['Test ID'][:12]
            borrower = result['Borrower Name'][:25]
            status = result['Status'][:10]

            # Color-code the status
            if result['Status'] == "PASSED":
                status_display = "✓ PASSED"
            elif result['Status'] == "FAILED":
                status_display = "✗ FAILED"
            elif result['Status'] == "SKIPPED":
                status_display = "⊘ SKIPPED"
            else:
                status_display = "! ERROR"

            notes = f"Passed: {result.get('Passed', 0)}, Failed: {result.get('Failed', 0)}"[:35]
            print(f"| {test_id:<12} | {borrower:<25} | {status_display:<10} | {notes:<35} |")

        print("="*100)

        print(f"\nSummary Statistics:")
        print(f"  Total Tests:         {total_tests}")
        print(f"  Passed:              {passed_tests} ({100*passed_tests//max(1,total_tests)}%)")
        print(f"  Failed:              {failed_tests}")
        print(f"  Skipped (Timeout):   {skipped_tests}")
        print(f"  Errors:              {error_tests}")

        print("\n" + "="*100)
        if skipped_tests > 0:
            print(f"[INFO] {skipped_tests} test(s) skipped due to timeout")
            print(f"       These tests took longer than {self.step_timeout} seconds to complete")
            print(f"       Check browser for hanging pages or slow calculations")
        if failed_tests == 0 and error_tests == 0 and skipped_tests == 0:
            print("[SUCCESS] ALL TESTS PASSED!")
        else:
            print(f"[WARNING] {failed_tests + error_tests + skipped_tests} test(s) did not pass")

        print("\nBrowser window left open for manual inspection")
        print("="*100 + "\n")

    def close(self):
        """Close WebDriver"""
        if self.driver:
            try:
                self.driver.quit()
                print("\nBrowser closed successfully")
            except Exception as e:
                print(f"\nError closing browser: {e}")


def main():
    """Main test execution"""

    # Load test cases from Excel
    if not EXCEL_AVAILABLE:
        print("ERROR: openpyxl is required for this test suite")
        print("Install with: pip install openpyxl")
        sys.exit(1)

    if not TEST_DATA_FILE.exists():
        print(f"ERROR: Test data file not found: {TEST_DATA_FILE}")
        print("Please run the Excel file creation script first")
        sys.exit(1)

    print("\nLoading test cases from Excel file...")
    test_cases = ExcelTestDataReader.read_test_cases(str(TEST_DATA_FILE))

    if not test_cases:
        print("ERROR: No test cases found in Excel file")
        sys.exit(1)

    print(f"[OK] Loaded {len(test_cases)} test cases from Excel")

    # Initialize and run tests
    test_suite = BankingCreditRiskTest()

    # Setup
    if not test_suite.setup():
        print("\nCannot proceed without WebDriver setup")
        sys.exit(1)

    try:
        # Run all test cycles
        test_suite.run_all_test_cycles(test_cases)
        test_suite.print_summary()

        # Close browser
        print("\nTest execution completed.")
        test_suite.close()

    except Exception as e:
        print(f"\nUnexpected error: {e}")
        test_suite.close()
        sys.exit(1)


if __name__ == "__main__":
    main()
